import cv2
import face_recognition
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def recognize_and_save_attendance(workbook):
    worksheet = workbook.active

    known_face_encodings = {}
    known_face_names = {}

    known_faces_dir = "known_faces"

    for person_dir in os.listdir(known_faces_dir):
        if os.path.isdir(os.path.join(known_faces_dir, person_dir)):
            known_face_encodings[person_dir] = []
            known_face_names[person_dir] = []

            for filename in os.listdir(os.path.join(known_faces_dir, person_dir)):
                if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                    image = face_recognition.load_image_file(os.path.join(known_faces_dir, person_dir, filename))
                    face_encodings = face_recognition.face_encodings(image)
                    if len(face_encodings) > 0:
                        known_face_encodings[person_dir].append(face_encodings[0])
                        known_face_names[person_dir].append(os.path.splitext(filename)[0])

    while True:
        video_capture = cv2.VideoCapture(0)

        while True:
            ret, frame = video_capture.read()

            face_locations = face_recognition.face_locations(frame)
            if not face_locations:
                continue  

            face_encodings = face_recognition.face_encodings(frame, face_locations)

            for face_encoding in face_encodings:
                for person_name, encodings in known_face_encodings.items():
                    matches = face_recognition.compare_faces(encodings, face_encoding, tolerance=0.6)
                    if True in matches:
                        name = person_name

                        current_date = datetime.now().strftime("%Y-%m-%d")
                        current_time = datetime.now().strftime("%H:%M:%S")

                        worksheet.append([name, current_date, current_time])

                        workbook.save("attendance.xlsx")

                        video_capture.release()
                        cv2.destroyAllWindows()
                        return

            for (top, right, bottom, left) in face_locations:
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, "Unknown", (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

            cv2.imshow("Attendance System", frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                video_capture.release()
                cv2.destroyAllWindows()
                break

if os.path.isfile("attendance.xlsx"):
    workbook = load_workbook("attendance.xlsx")
else:
    workbook = Workbook()
    workbook.active.title = "Attendance"

while True:
    recognize_and_save_attendance(workbook)