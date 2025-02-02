import cv2
import face_recognition
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import dlib

# Load face detection model from dlib
detector = dlib.get_frontal_face_detector()

# Function to perform liveness detection
def is_live(frame, face_location, eye_movement_threshold=3):
    # Extract the face from the frame
    top, right, bottom, left = face_location
    face_image = frame[top:bottom, left:right]

    # Convert the face image to grayscale for better liveness detection
    gray_face = cv2.cvtColor(face_image, cv2.COLOR_BGR2GRAY)

    # Use dlib to perform face detection
    faces = detector(gray_face)

    # If more than one face is detected, it's likely a photo or video
    if len(faces) != 1:
        return False

    # Detect eyes using face landmarks
    predictor_path = "shape_predictor_68_face_landmarks.dat"  # Download from http://dlib.net/files/shape_predictor_68_face_landmarks.dat.bz2
    predictor = dlib.shape_predictor(predictor_path)
    face_landmarks = predictor(gray_face, dlib.rectangle(0, 0, gray_face.shape[1], gray_face.shape[0]))

    # Extract eye landmarks (assuming 68 landmarks in total)
    left_eye_landmarks = list(range(36, 42))
    right_eye_landmarks = list(range(42, 48))

    # Calculate the movement in the eyes
    left_eye_movement = sum([abs(face_landmarks.part(i).x - face_landmarks.part(i-1).x) for i in left_eye_landmarks])
    right_eye_movement = sum([abs(face_landmarks.part(i).x - face_landmarks.part(i-1).x) for i in right_eye_landmarks])

    # Check if eye movement exceeds the threshold
    return left_eye_movement > eye_movement_threshold and right_eye_movement > eye_movement_threshold

# Function to recognize and save attendance with anti-spoofing and eye movement check
def recognize_and_save_attendance(workbook):
    worksheet = workbook.active

    known_face_encodings = {}
    known_face_names = {}

    # Replace "known_faces" with the path to your directory containing subdirectories for each person
    known_faces_dir = "known_faces"

    for person_dir in os.listdir(known_faces_dir):
        if os.path.isdir(os.path.join(known_faces_dir, person_dir)):
            known_face_encodings[person_dir] = []
            known_face_names[person_dir] = []

            for filename in os.listdir(os.path.join(known_faces_dir, person_dir)):
                if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                    image = face_recognition.load_image_file(os.path.join(known_faces_dir, person_dir, filename))
                    face_encodings = face_recognition.face_encodings(image)
                    if len(face_encodings) > 0:
                        known_face_encodings[person_dir].append(face_encodings[0])
                        known_face_names[person_dir].append(os.path.splitext(filename)[0])

    video_capture = cv2.VideoCapture(0)

    while True:
        ret, frame = video_capture.read()

        face_locations = face_recognition.face_locations(frame)
        if not face_locations:
            continue  # If no face is detected, continue with the next frame

        face_encodings = face_recognition.face_encodings(frame, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            for person_name, encodings in known_face_encodings.items():
                matches = face_recognition.compare_faces(encodings, face_encoding, tolerance=0.6)
                if True in matches and is_live(frame, face_location):
                    name = person_name

                    current_date = datetime.now().strftime("%Y-%m-%d")
                    current_time = datetime.now().strftime("%H:%M:%S")

                    worksheet.append([name, current_date, current_time])

                    # Save attendance
                    workbook.save("attendance.xlsx")

                    video_capture.release()
                    cv2.destroyAllWindows()
                    return

        for (top, right, bottom, left) in face_locations:
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, "Unknown", (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

        cv2.imshow("Attendance System", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            video_capture.release()
            cv2.destroyAllWindows()
            break

# Check if the attendance file exists, and load it if it does, else create a new one
if os.path.isfile("attendance.xlsx"):
    workbook = load_workbook("attendance.xlsx")
else:
    workbook = Workbook()
    workbook.active.title = "Attendance"

# Main loop to recognize and save attendance repeatedly
while True:
    recognize_and_save_attendance(workbook)